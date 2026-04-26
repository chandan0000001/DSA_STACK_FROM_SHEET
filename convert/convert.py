from openpyxl import load_workbook
import json

wb = load_workbook("questionSheet.xlsx")
ws = wb.active

data = []

for index, row in enumerate(ws.iter_rows(min_row=2), start=1):
    topic = row[0].value
    question_cell = row[1]
    difficulty = row[2].value

    name = question_cell.value
    link = question_cell.hyperlink.target if question_cell.hyperlink else ""

    # optional: clean link
    if link:
        link = link.split("?")[0]

    data.append({
        "name": name,
        "link": link,
        "tags": [topic] if topic else [],
        "difficulty": difficulty,
        "order": index   # 🔥 ADD THIS
    })

with open("questions.json", "w") as f:
    json.dump(data, f, indent=2)

print("✅ Done! JSON file created")